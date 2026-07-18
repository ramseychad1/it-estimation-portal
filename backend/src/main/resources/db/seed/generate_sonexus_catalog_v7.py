#!/usr/bin/env python3
"""Generate sonexus_catalog_v7.sql from the Sonexus V7 workbook.

Reads 'Production Baseline Catalog' (description + P25/Median/P75 sizing)
and 'Shore Mix by BRT & Phase' (authoritative category/team + per-phase
onshore/offshore hours) and emits an idempotent Postgres DO $$ block
matching the style of the old documents.local/prompts/seed_estimation_data.sql.

Run from the repo root. The source workbook lives in documents.local/
(gitignored — contains real client SOW $ amounts) and is not checked in;
re-derive it from a future workbook revision by pointing SRC at the new file.
Output lands in this same directory, which Flyway does NOT scan
(flyway.locations = classpath:db/migration only) — run it manually via psql,
it is never auto-executed.
"""
import openpyxl

SRC = "documents.local/estimation_files/Sonexus_Estimation_Catalog_Baseline_V7.xlsx"
OUT = "backend/src/main/resources/db/seed/sonexus_catalog_v7.sql"

PHASE_NAMES = [
    "Analysis", "Configuration", "Development", "Testing", "UAT",
    "Deployment", "Release Management", "Hypercare", "Monitoring", "Admin Activity",
]

EXCLUDED_CATEGORIES = {"Validation & Operational Support", "Re-Enrollment"}

# Reconcile the workbook's team spelling with the existing DB team.
TEAM_NAME_MAP = {"Telecom (EIT)": "Telecomm (EIT)"}
EXISTING_TEAMS = {
    "Salesforce", "Pixie", "Telecomm (EIT)", "OSP and DWH",
    "Pharmacy Team", "Full Stack Portals", "Data Hub", "CAPII",
}


def sql_str(v):
    if v is None:
        return "NULL"
    return "'" + str(v).replace("'", "''") + "'"


def sql_num(v):
    return "0" if v is None else repr(round(float(v), 2))


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)

    cat_ws = wb["Production Baseline Catalog"]
    cat_rows = list(cat_ws.iter_rows(values_only=True))[1:]
    catalog = {}
    for r in cat_rows:
        brt = r[1]
        if brt is None:
            continue
        catalog[brt] = dict(description=r[2], v6_samples=r[12], v6_small=r[14], v6_med=r[15], v6_large=r[16])

    shore_ws = wb["Shore Mix by BRT & Phase"]
    shore_rows = list(shore_ws.iter_rows(values_only=True))[2:-1]  # skip 2 header rows, skip grand total

    header1 = list(shore_ws.iter_rows(values_only=True))[0]
    for i, pname in enumerate(PHASE_NAMES):
        base = 9 + i * 3
        assert header1[base] == pname, f"column layout mismatch at {base}: {header1[base]!r} != {pname!r}"

    containers = {}  # category -> {team, subfeatures: [...]}
    for r in shore_rows:
        category, team, brt, n = r[0], r[1], r[2], r[3]
        if not n or n <= 0:
            continue
        if category in EXCLUDED_CATEGORIES:
            continue
        cat_info = catalog.get(brt)
        if cat_info is None:
            print(f"WARNING: {brt!r} has samples in Shore Mix but no Production Baseline Catalog row; skipping")
            continue

        row_total = r[4] or 0
        phases = []
        for i, pname in enumerate(PHASE_NAMES):
            base = 9 + i * 3
            chii, us = r[base] or 0, r[base + 1] or 0
            phase_total = chii + us
            phase_pct = (phase_total / row_total) if row_total else 0
            offshore_pct = (chii / phase_total) if phase_total else 0
            phases.append((pname, phase_pct, offshore_pct))

        db_team = TEAM_NAME_MAP.get(team, team)
        c = containers.setdefault(category, {"team": db_team, "subfeatures": []})
        if c["team"] != db_team:
            raise ValueError(f"Team conflict in category {category!r}: {c['team']!r} vs {db_team!r}")
        c["subfeatures"].append(dict(
            brt=brt, description=cat_info["description"],
            small=cat_info["v6_small"] or 0, med=cat_info["v6_med"] or 0, large=cat_info["v6_large"] or 0,
            phases=phases,
        ))

    new_teams = sorted({c["team"] for c in containers.values()} - EXISTING_TEAMS)

    n_containers = len(containers)
    n_subfeatures = sum(len(c["subfeatures"]) for c in containers.values())
    print(f"Containers: {n_containers}, SubFeatures: {n_subfeatures}, New teams: {new_teams}")

    lines = []
    lines.append("-- =========================================================================")
    lines.append("-- Sonexus V7 catalog seed — Products, SubFeatures, Estimate Templates")
    lines.append("--")
    lines.append("-- Generated from Sonexus_Estimation_Catalog_Baseline_V7.xlsx")
    lines.append("-- ('Production Baseline Catalog' + 'Shore Mix by BRT & Phase' tabs).")
    lines.append("-- Run manually via psql after V41__sonexus_phase_and_catalog_reset.sql has")
    lines.append("-- been applied. Idempotent — safe to re-run.")
    lines.append("-- =========================================================================")
    lines.append("")
    lines.append("DO $$")
    lines.append("DECLARE")
    lines.append("  v_admin_id BIGINT := 1;")
    lines.append("  v_team_id BIGINT;")
    lines.append("  v_product_id BIGINT;")
    lines.append("  v_sub_feature_id BIGINT;")
    lines.append("  v_template_id BIGINT;")
    lines.append("  v_phase_id BIGINT;")
    lines.append("BEGIN")
    lines.append("")
    lines.append("  -- ---- new teams ---------------------------------------------------------")
    for t in new_teams:
        lines.append(f"  INSERT INTO teams (name, description, active, created_by, updated_by)")
        lines.append(f"  SELECT {sql_str(t)}, {sql_str('Sonexus V7 catalog team')}, true, v_admin_id, v_admin_id")
        lines.append(f"  WHERE NOT EXISTS (SELECT 1 FROM teams WHERE LOWER(name) = LOWER({sql_str(t)}));")
        lines.append("")

    for category in sorted(containers.keys()):
        c = containers[category]
        lines.append(f"  -- =========================================================================")
        lines.append(f"  -- Container: {category}  (Team: {c['team']})")
        lines.append(f"  -- =========================================================================")
        lines.append(f"  SELECT id INTO v_team_id FROM teams WHERE LOWER(name) = LOWER({sql_str(c['team'])});")
        lines.append("")
        lines.append("  INSERT INTO products (name, description, mode, active, team_id, created_by, updated_by)")
        lines.append(f"  VALUES ({sql_str(category)}, NULL, 'CONTAINER', true, v_team_id, v_admin_id, v_admin_id)")
        lines.append("  ON CONFLICT DO NOTHING;")
        lines.append(f"  SELECT id INTO v_product_id FROM products WHERE LOWER(name) = LOWER({sql_str(category)});")
        lines.append("")

        for sf in c["subfeatures"]:
            lines.append(f"  -- SubFeature: {sf['brt']}")
            lines.append(f"  -- Small={sf['small']}  Med={sf['med']}  Large={sf['large']} hrs (blended)")
            lines.append("  INSERT INTO sub_features (product_id, name, description, active, created_by, updated_by)")
            lines.append(f"  VALUES (v_product_id, {sql_str(sf['brt'])}, {sql_str(sf['description'])}, true, v_admin_id, v_admin_id)")
            lines.append("  ON CONFLICT DO NOTHING;")
            lines.append("  SELECT id INTO v_sub_feature_id FROM sub_features")
            lines.append(f"    WHERE product_id = v_product_id AND LOWER(name) = LOWER({sql_str(sf['brt'])});")
            lines.append("")
            lines.append("  INSERT INTO estimate_templates (sub_feature_id, version_number, is_active, created_by)")
            lines.append("  VALUES (v_sub_feature_id, 1, true, v_admin_id);")
            lines.append("  v_template_id := currval('estimate_templates_id_seq');")
            lines.append("")

            for pname, phase_pct, offshore_pct in sf["phases"]:
                onshore_low = sf["small"] * phase_pct * (1 - offshore_pct)
                onshore_med = sf["med"] * phase_pct * (1 - offshore_pct)
                onshore_high = sf["large"] * phase_pct * (1 - offshore_pct)
                offshore_low = sf["small"] * phase_pct * offshore_pct
                offshore_med = sf["med"] * phase_pct * offshore_pct
                offshore_high = sf["large"] * phase_pct * offshore_pct

                lines.append(f"  SELECT id INTO v_phase_id FROM sdlc_phases WHERE LOWER(name) = LOWER({sql_str(pname)});")
                lines.append("  INSERT INTO estimate_template_lines (template_id, sdlc_phase_id,")
                lines.append("    onshore_low, onshore_med, onshore_high, offshore_low, offshore_med, offshore_high)")
                lines.append(f"  VALUES (v_template_id, v_phase_id,")
                lines.append(f"    {sql_num(onshore_low)}, {sql_num(onshore_med)}, {sql_num(onshore_high)}, "
                              f"{sql_num(offshore_low)}, {sql_num(offshore_med)}, {sql_num(offshore_high)});")
            lines.append("")

    lines.append("END $$;")
    lines.append("")

    with open(OUT, "w") as f:
        f.write("\n".join(lines))
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
